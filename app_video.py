import os
import re
import io
import base64
import datetime
import tempfile
import threading
import sys
import types
import platform
from concurrent.futures import ThreadPoolExecutor, as_completed
from streamlit.runtime.scriptrunner import add_script_run_ctx, get_script_run_ctx

# ─────────────────────────────────────────────
# Streamlit Cloud/Linux 호환 처리
# ─────────────────────────────────────────────
# core_video.py 안에 Windows 전용 모듈(win32com.client)을 import하는 코드가 있으면
# Streamlit Cloud(Linux)에서는 앱 시작 단계에서 ModuleNotFoundError가 발생합니다.
# 영상 분석 기능은 win32com이 없어도 동작할 수 있으므로, 비-Windows 환경에서는
# 더미 모듈을 먼저 등록해 core_video import 자체가 실패하지 않도록 합니다.
if platform.system() != "Windows":
    if "win32com" not in sys.modules:
        win32com_stub = types.ModuleType("win32com")
        client_stub = types.ModuleType("win32com.client")

        def _dispatch_unavailable(*args, **kwargs):
            raise RuntimeError(
                "win32com.client는 Windows 전용 기능입니다. "
                "Streamlit Cloud에서는 Office COM 자동화/PPT 변환 기능을 사용할 수 없습니다."
            )

        client_stub.Dispatch = _dispatch_unavailable
        win32com_stub.client = client_stub
        sys.modules["win32com"] = win32com_stub
        sys.modules["win32com.client"] = client_stub

    # 일부 Windows 자동화 코드가 pythoncom을 함께 import하는 경우를 대비합니다.
    if "pythoncom" not in sys.modules:
        pythoncom_stub = types.ModuleType("pythoncom")
        pythoncom_stub.CoInitialize = lambda *args, **kwargs: None
        pythoncom_stub.CoUninitialize = lambda *args, **kwargs: None
        sys.modules["pythoncom"] = pythoncom_stub

import streamlit as st
import pandas as pd
from dotenv import load_dotenv
from PIL import Image as PILImage
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from core_video import (
    extract_audio,
    transcribe_audio,
    spell_check_segments,
    extract_and_filter_frames,
    spell_check_frames,
    attach_frames_to_audio_results,
    run_pipeline_parallel,
    set_encode_quality,
)

# ─────────────────────────────────────────────
# 페이지 설정
# ─────────────────────────────────────────────
st.set_page_config(page_title="AI 품질관리 시스템(콘텐츠)", page_icon="🎥", layout="wide")

st.title("🎥 AI 품질관리 시스템(콘텐츠)")
st.markdown(
    "MP4 영상을 업로드하면 AI가 **음성 대본(STT)** 및 **화면 속 텍스트(OCR)**를 분석하여 "
    "맞춤법·오타·띄어쓰기 오류를 잡아줍니다."
)

# ─────────────────────────────────────────────
# 공통 CSS
# ─────────────────────────────────────────────
st.markdown("""
<style>
.card {
    background: #fff;
    border: 1px solid #e0e4ec;
    border-radius: 10px;
    padding: 16px 20px;
    margin-bottom: 14px;
    font-family: 'Malgun Gothic', 'Apple SD Gothic Neo', sans-serif;
    box-shadow: 0 1px 4px rgba(0,0,0,0.06);
}
.card-header {
    display: flex;
    align-items: center;
    gap: 10px;
    margin-bottom: 12px;
    font-size: 13px;
    color: #666;
    border-bottom: 1px solid #f0f0f0;
    padding-bottom: 8px;
}
.badge-audio  { background:#4f8ef7; color:white; border-radius:5px; padding:2px 9px; font-size:12px; font-weight:bold; }
.badge-screen { background:#22b07d; color:white; border-radius:5px; padding:2px 9px; font-size:12px; font-weight:bold; }
.timestamp    { background:#f3f4f6; border-radius:4px; padding:2px 8px; font-size:12px; color:#555; font-family:monospace; }
.label-before { font-size:12px; font-weight:bold; color:#999; margin-bottom:4px; }
.label-after  { font-size:12px; font-weight:bold; color:#4f8ef7; margin-bottom:4px; }
.text-before  { background:#fff8f8; border-left:3px solid #f87171; border-radius:4px; padding:8px 12px; font-size:15px; color:#333; line-height:1.7; margin-bottom:8px; }
.text-after   { background:#f0fdf4; border-left:3px solid #34d399; border-radius:4px; padding:8px 12px; font-size:15px; color:#333; line-height:1.7; margin-bottom:6px; }
.reason-box   { font-size:12px; color:#888; margin-top:4px; }
.no-result    { text-align:center; padding:40px; color:#aaa; font-size:15px; }
.info-box     { background:#f8faff; border:1px solid #d0dcf5; border-radius:10px; padding:18px 22px; margin-bottom:20px; }
[data-testid="stSidebar"] hr { margin-top: 0.4rem !important; margin-bottom: 0.4rem !important; }
.info-box h4  { margin:0 0 12px 0; color:#3a5bbf; font-size:15px; }
.info-grid    { display:grid; grid-template-columns:1fr 1fr 1fr; gap:12px; }
.info-item label { font-size:11px; color:#888; display:block; margin-bottom:3px; }
.info-item span  { font-size:14px; font-weight:bold; color:#222; }
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────
# 엑셀 생성 함수
# ─────────────────────────────────────────────
def _b64_to_xl_image(b64_str, thumb_w=240, thumb_h=135):
    """Base64 JPEG 문자열을 openpyxl용 XLImage 객체로 변환합니다."""
    raw = base64.b64decode(b64_str)
    pil = PILImage.open(io.BytesIO(raw)).convert("RGB")
    pil.thumbnail((thumb_w, thumb_h), PILImage.LANCZOS)
    buf = io.BytesIO()
    pil.save(buf, format="JPEG", quality=85)
    buf.seek(0)
    return XLImage(buf)


SCORE_AUDIO_WEIGHT = 1.0
SCORE_SCREEN_WEIGHT = 1.4
SCORE_CURVE_BASE = 55.0
SCORE_MAX_DEDUCTION = 75.0


def calculate_score(audio_results, screen_results):
    """오류 건수가 많아도 급격히 0점으로 떨어지지 않는 완만한 품질 점수를 계산합니다."""
    audio_count = len(audio_results or [])
    screen_count = len(screen_results or [])
    weighted_errors = audio_count * SCORE_AUDIO_WEIGHT + screen_count * SCORE_SCREEN_WEIGHT
    deduction = SCORE_MAX_DEDUCTION * (weighted_errors / (weighted_errors + SCORE_CURVE_BASE)) if weighted_errors else 0
    score = round(max(0, 100 - deduction))
    if score >= 95:
        grade = "우수"
    elif score >= 85:
        grade = "양호"
    elif score >= 70:
        grade = "주의"
    else:
        grade = "개선 필요"
    return {
        "score": score,
        "grade": grade,
        "audio_count": audio_count,
        "screen_count": screen_count,
        "total_count": audio_count + screen_count,
        "deduction": round(deduction, 1),
        "weighted_errors": round(weighted_errors, 1),
    }


def build_score_rows(all_video_results):
    rows = []
    for vr in all_video_results or []:
        rows.append({
            "video": vr.get("name") or vr.get("filename") or "-",
            **calculate_score(vr.get("audio", []), vr.get("screen", [])),
        })
    return rows


def average_score(score_rows):
    if not score_rows:
        return 100
    return round(sum(r["score"] for r in score_rows) / len(score_rows), 1)


def render_score_summary_card(score_rows):
    if not score_rows:
        return

    overall_score = average_score(score_rows)
    total_audio = sum(r["audio_count"] for r in score_rows)
    total_screen = sum(r["screen_count"] for r in score_rows)
    total_items = sum(r["total_count"] for r in score_rows)
    grade = (
        "우수" if overall_score >= 95 else
        "양호" if overall_score >= 85 else
        "주의" if overall_score >= 70 else
        "개선 필요"
    )
    bar_width = max(0, min(100, overall_score))
    score_display = f"{overall_score:g}"

    st.markdown("### 🏅 문서 품질 점수")
    st.markdown(
        f"""
        <div style="
            background:linear-gradient(135deg,#161b2f 0%,#11182b 52%,#17243f 100%);
            border:1px solid rgba(93,135,213,.25);
            border-radius:10px;
            box-shadow:0 12px 28px rgba(17,24,39,.22);
            padding:22px 24px 20px;
            margin:6px 0 18px;
            color:#e8eefc;
            font-family:'Malgun Gothic','Apple SD Gothic Neo',sans-serif;">
            <div style="font-size:13px;font-weight:700;color:#dbe7ff;margin-bottom:14px;">
                📊 문서 품질 점수 <span style="font-weight:500;color:#91a2c8;">(오류 건수 기반)</span>
            </div>
            <div style="display:flex;align-items:center;gap:18px;margin-bottom:10px;">
                <div style="min-width:92px;">
                    <div style="font-size:44px;line-height:46px;font-weight:800;color:#34a8ff;white-space:nowrap;">{score_display}</div>
                    <div style="font-size:12px;color:#8fa1c6;margin-top:2px;">/ 100점</div>
                </div>
                <div style="display:flex;align-items:center;gap:8px;margin-top:-14px;">
                    <span style="width:12px;height:12px;border-radius:50%;background:#2f9dff;display:inline-block;"></span>
                    <span style="font-size:15px;font-weight:700;color:#dbeafe;">{grade}</span>
                </div>
            </div>
            <div style="margin:0 0 18px 92px;">
                <div style="height:9px;border-radius:999px;background:#263149;overflow:hidden;">
                    <div style="width:{bar_width}%;height:100%;border-radius:999px;background:linear-gradient(90deg,#2f9dff,#38bdf8);"></div>
                </div>
                <div style="display:flex;justify-content:space-between;font-size:10px;color:#66779d;margin-top:5px;">
                    <span>0</span><span>50</span><span>100</span>
                </div>
            </div>
            <div style="display:grid;grid-template-columns:repeat(3,minmax(0,1fr));gap:8px;">
                <div style="background:rgba(255,255,255,.055);border:1px solid rgba(255,255,255,.08);border-radius:7px;padding:12px;text-align:center;">
                    <div style="font-size:11px;color:#9fb0d2;margin-bottom:6px;">총 오류 수</div>
                    <div style="font-size:19px;font-weight:800;color:#ffffff;">{total_items}건</div>
                </div>
                <div style="background:rgba(255,255,255,.055);border:1px solid rgba(255,255,255,.08);border-radius:7px;padding:12px;text-align:center;">
                    <div style="font-size:11px;color:#9fb0d2;margin-bottom:6px;">음성 대본</div>
                    <div style="font-size:19px;font-weight:800;color:#ff6b6b;">{total_audio}건</div>
                </div>
                <div style="background:rgba(255,255,255,.055);border:1px solid rgba(255,255,255,.08);border-radius:7px;padding:12px;text-align:center;">
                    <div style="font-size:11px;color:#9fb0d2;margin-bottom:6px;">화면 텍스트</div>
                    <div style="font-size:19px;font-weight:800;color:#f8c14a;">{total_screen}건</div>
                </div>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def make_excel(audio_results, screen_results, reviewer, lesson_name, lesson_num, review_date=None):
    wb = Workbook()

    def strip_html(text):
        return re.sub(r'<[^>]+>', '', text or '')

    # ── 공통 스타일 ───────────────────────────────
    header_fill  = PatternFill("solid", start_color="1E2533")
    audio_fill   = PatternFill("solid", start_color="EBF3FF")
    screen_fill  = PatternFill("solid", start_color="EAFAF3")
    all_fill     = PatternFill("solid", start_color="F5F7FA")
    meta_fill    = PatternFill("solid", start_color="F0F4FF")
    white_fill   = PatternFill("solid", start_color="FFFFFF")

    header_font  = Font(name="맑은 고딕", bold=True, color="FFFFFF",  size=11)
    title_font   = Font(name="맑은 고딕", bold=True, color="1E2533",  size=16)
    body_font    = Font(name="맑은 고딕",             color="111111",  size=10)
    meta_label_f = Font(name="맑은 고딕", bold=True,  color="888888",  size=9)
    meta_value_f = Font(name="맑은 고딕", bold=True,  color="1E2533",  size=11)

    thin   = Side(style="thin",   color="CCCCCC")
    medium = Side(style="medium", color="AAAAAA")
    thin_b   = Border(left=thin,   right=thin,   top=thin,   bottom=thin)
    thick_b  = Border(left=medium, right=medium, top=medium, bottom=medium)
    center   = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_top = Alignment(horizontal="left",   vertical="top",    wrap_text=True)

    IMG_ROW_H = 105   # 이미지 행 높이 (포인트, ≈ 140px)
    IMG_COL   = 6     # 이미지 컬럼 번호 (F) — 조치결과(G) 앞
    IMG_COL_W = 36    # 이미지 컬럼 너비 (문자 단위)
    action_fill = PatternFill("solid", start_color="FFFDE7")  # 연노랑 — 사용자 입력 유도

    def write_meta(ws, has_image_col=False):
        """제목 + 메타(검토자/차시명/차시) 블록을 씁니다."""
        last_col = "G" if has_image_col else "F"

        ws.merge_cells(f"A1:{last_col}1")
        ws["A1"] = "영상 맞춤법 교정 결과 보고서"
        ws["A1"].font      = title_font
        ws["A1"].fill      = white_fill
        ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 36

        ws.row_dimensions[2].height = 6
        date_str = review_date.strftime("%Y년 %m월 %d일") if review_date else "-"
        for i, (lbl, val) in enumerate([
            ("검토자",  reviewer    or "-"),
            ("차시명",  lesson_name or "-"),
            ("차시",    lesson_num  or "-"),
            ("검토일자", date_str),
        ]):
            r = 3 + i
            ws.merge_cells(f"A{r}:B{r}")
            ws.merge_cells(f"C{r}:{last_col}{r}")
            ws[f"A{r}"] = lbl;  ws[f"A{r}"].font = meta_label_f
            ws[f"A{r}"].fill = meta_fill; ws[f"A{r}"].alignment = center
            ws[f"A{r}"].border = thin_b
            ws[f"C{r}"] = val;  ws[f"C{r}"].font = meta_value_f
            ws[f"C{r}"].fill = meta_fill; ws[f"C{r}"].alignment = left_top
            ws[f"C{r}"].border = thin_b
            ws.row_dimensions[r].height = 22
        ws.row_dimensions[7].height = 8  # 여백 (메타 4개 → 3~6행, 여백 7행)

    def write_headers(ws, has_image_col=False):
        """컬럼 헤더 행(8번)을 씁니다."""
        if has_image_col:
            headers = ["구분", "시간", "수정 전", "수정 후", "교정 사유", "수정 전 화면 이미지", "조치결과", "수정 후 화면 이미지"]
        else:
            headers = ["구분", "시간", "수정 전", "수정 후", "교정 사유", "조치결과"]
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=8, column=col, value=h)
            c.font = header_font; c.fill = header_fill
            c.alignment = center; c.border = thick_b
        ws.row_dimensions[8].height = 24

    def write_rows(ws, results, badge_label, row_fill, has_image_col=False):
        """데이터 행을 쓰고 이미지가 있으면 삽입합니다."""
        total_cols = 8 if has_image_col else 6
        if not results:
            span = f"A9:{get_column_letter(total_cols)}9"
            ws.merge_cells(span)
            ws["A9"] = "교정 항목 없음"
            ws["A9"].alignment = center; ws["A9"].font = body_font
            ws.row_dimensions[9].height = 22
            return

        for i, r in enumerate(results):
            row_idx = 9 + i
            # ★ v3: 화면 시트에서는 각 행의 issue_type(좌측상단/STEP 로고/화면 텍스트/맞춤법)을 구분 값으로
            if badge_label == "화면 텍스트":
                gbn = r.get("_issue_type") or r.get("구분") or badge_label
            else:
                gbn = badge_label if badge_label else r.get("구분", "")
            fill = row_fill if i % 2 == 0 else white_fill
            vals = [
                gbn,
                r.get("시간", ""),
                strip_html(r.get("수정 전", "")),
                strip_html(r.get("수정 후", "")),
                r.get("교정 사유", ""),
            ]
            for col, v in enumerate(vals, 1):
                c = ws.cell(row=row_idx, column=col, value=v)
                c.font = body_font; c.fill = fill; c.border = thin_b
                c.alignment = center if col <= 2 else left_top

            # 이미지 삽입 (col 6)
            # 조치결과 칸은 이미지 뒤 col 7에 위치
            if has_image_col:
                b64 = r.get("image_b64", "")
                img_cell = ws.cell(row=row_idx, column=IMG_COL)
                img_cell.fill = fill; img_cell.border = thin_b
                if b64:
                    try:
                        xl_img = _b64_to_xl_image(b64)
                        xl_img.anchor = f"{get_column_letter(IMG_COL)}{row_idx}"
                        ws.add_image(xl_img)
                    except Exception as e:
                        img_cell.value = "이미지 오류"
                ws.row_dimensions[row_idx].height = IMG_ROW_H
                # 조치결과 칸 (col 7) — 연노랑 빈칸
                ac = ws.cell(row=row_idx, column=7, value="")
                ac.fill = action_fill; ac.border = thin_b; ac.alignment = center
                # 수정 후 화면 이미지 칸 (col 8) — 연노랑 빈칸, 사용자가 직접 삽입
                af = ws.cell(row=row_idx, column=8, value="")
                af.fill = action_fill; af.border = thin_b; af.alignment = center
            else:
                ws.row_dimensions[row_idx].height = 40

    def set_col_widths(ws, has_image_col=False):
        # 구분, 시간, 수정전, 수정후, 교정사유[, 수정전이미지, 조치결과, 수정후이미지]
        widths = [12, 12, 34, 34, 28]
        if has_image_col:
            widths += [IMG_COL_W, 9, IMG_COL_W]
        for col, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = w

    # ── 시트 1: 음성 대본 (이미지 포함) ─────────────
    ws_audio = wb.active
    ws_audio.title = "음성 대본"
    write_meta(ws_audio, has_image_col=True)
    write_headers(ws_audio, has_image_col=True)
    write_rows(ws_audio, audio_results, "음성 대본", audio_fill, has_image_col=True)
    set_col_widths(ws_audio, has_image_col=True)

    # ── 시트 2: 화면 텍스트 (이미지 포함) ───────────
    ws_screen = wb.create_sheet("화면 텍스트")
    write_meta(ws_screen, has_image_col=True)
    write_headers(ws_screen, has_image_col=True)
    write_rows(ws_screen, screen_results, "화면 텍스트", screen_fill, has_image_col=True)
    set_col_widths(ws_screen, has_image_col=True)

    # ── 시트 3: 통합 (시간순, 이미지 포함) ──────────
    all_results = sorted(
        list(audio_results) + list(screen_results),
        key=lambda x: x.get("시간", "")
    )
    ws_all = wb.create_sheet("통합")
    write_meta(ws_all, has_image_col=True)
    write_headers(ws_all, has_image_col=True)
    write_rows(ws_all, all_results, "", all_fill, has_image_col=True)
    set_col_widths(ws_all, has_image_col=True)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def make_excel_combined(all_video_results, reviewer, lesson_name, lesson_num, review_date=None):
    """
    여러 영상의 결과를 하나의 엑셀 파일로 통합합니다.
    구조: [영상명_음성], [영상명_화면], ... + [📋 전체통합] 시트
    """
    wb = Workbook()
    first = True  # 첫 번째 시트는 active 시트 사용

    def strip_html(text):
        return re.sub(r'<[^>]+>', '', text or '')

    # 공통 스타일
    header_fill  = PatternFill("solid", start_color="1E2533")
    audio_fill   = PatternFill("solid", start_color="EBF3FF")
    screen_fill  = PatternFill("solid", start_color="EAFAF3")
    all_fill     = PatternFill("solid", start_color="F5F7FA")
    meta_fill    = PatternFill("solid", start_color="F0F4FF")
    white_fill   = PatternFill("solid", start_color="FFFFFF")
    action_fill  = PatternFill("solid", start_color="FFFDE7")

    header_font  = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=11)
    title_font   = Font(name="맑은 고딕", bold=True, color="1E2533", size=16)
    body_font    = Font(name="맑은 고딕", color="111111", size=10)
    meta_label_f = Font(name="맑은 고딕", bold=True, color="888888", size=9)
    meta_value_f = Font(name="맑은 고딕", bold=True, color="1E2533", size=11)
    video_title_font = Font(name="맑은 고딕", bold=True, color="3A5BBF", size=13)

    thin   = Side(style="thin",   color="CCCCCC")
    medium = Side(style="medium", color="AAAAAA")
    thin_b  = Border(left=thin,   right=thin,   top=thin,   bottom=thin)
    thick_b = Border(left=medium, right=medium, top=medium, bottom=medium)
    center   = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_top = Alignment(horizontal="left",   vertical="top",    wrap_text=True)

    IMG_ROW_H = 105
    IMG_COL   = 6
    IMG_COL_W = 36

    def _write_sheet(ws, results, badge_label, row_fill, video_name, has_img=True):
        """한 시트에 메타 + 헤더 + 데이터를 씁니다."""
        last_col = "H" if has_img else "F"
        date_str = review_date.strftime("%Y년 %m월 %d일") if review_date else "-"

        # 제목
        ws.merge_cells(f"A1:{last_col}1")
        ws["A1"] = "영상 맞춤법 교정 결과 보고서"
        ws["A1"].font = title_font; ws["A1"].fill = white_fill
        ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 36

        # 영상명 표시 (통합본에서 어느 영상인지 구분)
        ws.merge_cells(f"A2:{last_col}2")
        ws["A2"] = f"📹 {video_name}"
        ws["A2"].font = video_title_font
        ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[2].height = 22

        # 메타 (행 3~6)
        ws.row_dimensions[3].height = 4  # 여백
        for i, (lbl, val) in enumerate([
            ("검토자",  reviewer    or "-"),
            ("차시명",  lesson_name or "-"),
            ("차시",    lesson_num  or "-"),
            ("검토일자", date_str),
        ]):
            r = 4 + i
            ws.merge_cells(f"A{r}:B{r}")
            ws.merge_cells(f"C{r}:{last_col}{r}")
            ws[f"A{r}"] = lbl; ws[f"A{r}"].font = meta_label_f
            ws[f"A{r}"].fill = meta_fill; ws[f"A{r}"].alignment = center
            ws[f"A{r}"].border = thin_b
            ws[f"C{r}"] = val; ws[f"C{r}"].font = meta_value_f
            ws[f"C{r}"].fill = meta_fill; ws[f"C{r}"].alignment = left_top
            ws[f"C{r}"].border = thin_b
            ws.row_dimensions[r].height = 22
        ws.row_dimensions[8].height = 6  # 여백

        # 헤더 (행 9)
        if has_img:
            headers = ["구분", "시간", "수정 전", "수정 후", "교정 사유",
                       "수정 전 화면 이미지", "조치결과", "수정 후 화면 이미지"]
        else:
            headers = ["구분", "시간", "수정 전", "수정 후", "교정 사유", "조치결과"]
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=9, column=col, value=h)
            c.font = header_font; c.fill = header_fill
            c.alignment = center; c.border = thick_b
        ws.row_dimensions[9].height = 24

        # 데이터 (행 10~)
        total_cols = 8 if has_img else 6
        if not results:
            span = f"A10:{get_column_letter(total_cols)}10"
            ws.merge_cells(span)
            ws["A10"] = "교정 항목 없음"
            ws["A10"].alignment = center; ws["A10"].font = body_font
            ws.row_dimensions[10].height = 22
        else:
            for i, r in enumerate(results):
                row_idx = 10 + i
                # ★ v3: 화면 시트는 각 행의 issue_type(좌측상단/STEP 로고/화면 텍스트/맞춤법)을 구분 값으로
                if badge_label == "화면 텍스트":
                    gbn = r.get("_issue_type") or r.get("구분") or badge_label
                else:
                    gbn = badge_label if badge_label else r.get("구분", "")
                fill = row_fill if i % 2 == 0 else white_fill
                vals = [gbn, r.get("시간",""), strip_html(r.get("수정 전","")),
                        strip_html(r.get("수정 후","")), r.get("교정 사유","")]
                for col, v in enumerate(vals, 1):
                    c = ws.cell(row=row_idx, column=col, value=v)
                    c.font = body_font; c.fill = fill; c.border = thin_b
                    c.alignment = center if col <= 2 else left_top
                if has_img:
                    b64 = r.get("image_b64","")
                    img_cell = ws.cell(row=row_idx, column=IMG_COL)
                    img_cell.fill = fill; img_cell.border = thin_b
                    if b64:
                        try:
                            xl_img = _b64_to_xl_image(b64)
                            xl_img.anchor = f"{get_column_letter(IMG_COL)}{row_idx}"
                            ws.add_image(xl_img)
                        except Exception:
                            img_cell.value = "이미지 오류"
                    ws.row_dimensions[row_idx].height = IMG_ROW_H
                    ac = ws.cell(row=row_idx, column=7, value="")
                    ac.fill = action_fill; ac.border = thin_b; ac.alignment = center
                    af = ws.cell(row=row_idx, column=8, value="")
                    af.fill = action_fill; af.border = thin_b; af.alignment = center
                else:
                    ws.row_dimensions[row_idx].height = 40

        # 컬럼 너비
        widths = [12, 12, 34, 34, 28]
        if has_img:
            widths += [IMG_COL_W, 9, IMG_COL_W]
        else:
            widths += [18]
        for col, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = w

    # ── 영상별 시트 생성 ─────────────────────────────
    all_combined = []   # 전체통합 시트용
    for vr in all_video_results:
        vname = vr["name"]
        audio  = vr["audio"]
        screen = vr["screen"]

        # 시트명 최대 31자 제한 (엑셀 규칙)
        short = vname[:12] if len(vname) > 12 else vname

        # 음성 시트
        ws_a = wb.active if first else wb.create_sheet(f"{short}_음성")
        if first:
            ws_a.title = f"{short}_음성"
            first = False
        _write_sheet(ws_a, audio, "음성 대본", audio_fill, vname)

        # 화면 시트
        ws_s = wb.create_sheet(f"{short}_화면")
        _write_sheet(ws_s, screen, "화면 텍스트", screen_fill, vname)

        # 전체통합용 데이터에 영상명 추가
        for r in audio + screen:
            all_combined.append(dict(r, _video=vname))

    # ── 전체 통합 시트 ───────────────────────────────
    all_combined.sort(key=lambda x: (x.get("_video",""), x.get("시간","")))
    ws_all = wb.create_sheet("📋 전체통합")

    # 전체통합은 영상명 컬럼을 맨 앞에 추가
    last_col_all = "I"
    date_str_all = review_date.strftime("%Y년 %m월 %d일") if review_date else "-"
    ws_all.merge_cells(f"A1:{last_col_all}1")
    ws_all["A1"] = "영상 맞춤법 교정 결과 보고서 — 전체 통합"
    ws_all["A1"].font = title_font; ws_all["A1"].fill = white_fill
    ws_all["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws_all.row_dimensions[1].height = 36
    ws_all.row_dimensions[2].height = 4
    for i, (lbl, val) in enumerate([
        ("검토자",  reviewer    or "-"),
        ("차시명",  lesson_name or "-"),
        ("차시",    lesson_num  or "-"),
        ("검토일자", date_str_all),
    ]):
        r = 3 + i
        ws_all.merge_cells(f"A{r}:B{r}")
        ws_all.merge_cells(f"C{r}:{last_col_all}{r}")
        ws_all[f"A{r}"] = lbl; ws_all[f"A{r}"].font = meta_label_f
        ws_all[f"A{r}"].fill = meta_fill; ws_all[f"A{r}"].alignment = center
        ws_all[f"A{r}"].border = thin_b
        ws_all[f"C{r}"] = val; ws_all[f"C{r}"].font = meta_value_f
        ws_all[f"C{r}"].fill = meta_fill; ws_all[f"C{r}"].alignment = left_top
        ws_all[f"C{r}"].border = thin_b
        ws_all.row_dimensions[r].height = 22
    ws_all.row_dimensions[7].height = 6

    # 전체통합 헤더 (영상명 컬럼 추가)
    all_headers = ["영상명", "구분", "시간", "수정 전", "수정 후",
                   "교정 사유", "수정 전 화면 이미지", "조치결과", "수정 후 화면 이미지"]
    for col, h in enumerate(all_headers, 1):
        c = ws_all.cell(row=8, column=col, value=h)
        c.font = header_font; c.fill = header_fill
        c.alignment = center; c.border = thick_b
    ws_all.row_dimensions[8].height = 24

    if not all_combined:
        ws_all.merge_cells("A9:I9")
        ws_all["A9"] = "교정 항목 없음"
        ws_all["A9"].alignment = center; ws_all["A9"].font = body_font
        ws_all.row_dimensions[9].height = 22
    else:
        for i, r in enumerate(all_combined):
            row_idx = 9 + i
            fill = all_fill if i % 2 == 0 else white_fill
            vals = [
                r.get("_video",""),
                r.get("구분",""),
                r.get("시간",""),
                strip_html(r.get("수정 전","")),
                strip_html(r.get("수정 후","")),
                r.get("교정 사유",""),
            ]
            for col, v in enumerate(vals, 1):
                c = ws_all.cell(row=row_idx, column=col, value=v)
                c.font = body_font; c.fill = fill; c.border = thin_b
                c.alignment = center if col <= 3 else left_top
            # 이미지 (col 7)
            b64 = r.get("image_b64","")
            img_cell = ws_all.cell(row=row_idx, column=7)
            img_cell.fill = fill; img_cell.border = thin_b
            if b64:
                try:
                    xl_img = _b64_to_xl_image(b64)
                    xl_img.anchor = f"G{row_idx}"
                    ws_all.add_image(xl_img)
                except Exception:
                    img_cell.value = "이미지 오류"
            ws_all.row_dimensions[row_idx].height = IMG_ROW_H
            ac = ws_all.cell(row=row_idx, column=8, value="")
            ac.fill = action_fill; ac.border = thin_b; ac.alignment = center
            af = ws_all.cell(row=row_idx, column=9, value="")
            af.fill = action_fill; af.border = thin_b; af.alignment = center

    # 전체통합 컬럼 너비
    for col, w in enumerate([20, 12, 12, 34, 34, 28, IMG_COL_W, 9, IMG_COL_W], 1):
        ws_all.column_dimensions[get_column_letter(col)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ─────────────────────────────────────────────
# 결과 카드 렌더링
# ─────────────────────────────────────────────
def render_result_cards(results, badge_type="audio"):
    if not results:
        st.markdown('<div class="no-result">✅ 교정이 필요한 항목이 없습니다.</div>', unsafe_allow_html=True)
        return
    # ★ v3: 화면 이슈는 issue_type별 세부 뱃지 표시
    _issue_color = {
        "좌측상단":    "#8b5cf6",   # 보라
        "STEP 로고":   "#f97316",   # 주황
        "화면 텍스트": "#22b07d",   # 초록
        "맞춤법":     "#0ea5e9",   # 하늘
    }
    default_class = "badge-audio" if badge_type == "audio" else "badge-screen"
    default_label = "음성 대본" if badge_type == "audio" else "화면 텍스트"

    for i, r in enumerate(results, 1):
        reason_html = f'<div class="reason-box">📌 {r.get("교정 사유","")}</div>' if r.get("교정 사유") else ""
        # 세부 issue_type 우선 사용 (화면 이슈만 해당)
        it = r.get("_issue_type") or r.get("구분") or default_label
        if badge_type == "screen" and it in _issue_color:
            color = _issue_color[it]
            badge_html = (
                f'<span style="background:{color};color:white;border-radius:5px;'
                f'padding:2px 9px;font-size:12px;font-weight:bold;">{it}</span>'
            )
        else:
            badge_html = f'<span class="{default_class}">{default_label}</span>'

        st.markdown(f"""
        <div class="card">
            <div class="card-header">
                {badge_html}
                <span class="timestamp">{r.get("시간","")}</span>
                <span style="margin-left:auto;color:#bbb;font-size:12px;">#{i}</span>
            </div>
            <div class="label-before">수정 전</div>
            <div class="text-before">{r.get("수정 전","")}</div>
            <div class="label-after">수정 후</div>
            <div class="text-after">{r.get("수정 후","")}</div>
            {reason_html}
        </div>
        """, unsafe_allow_html=True)


# ─────────────────────────────────────────────
# API 키 설정 — .env / Streamlit Secrets 자동 로드
# ─────────────────────────────────────────────
# 로컬 실행: .env 파일의 OPENAI_API_KEY 사용
# Streamlit Cloud: Secrets의 OPENAI_API_KEY 사용 권장
load_dotenv()

api_key = os.getenv("OPENAI_API_KEY", "")

# Streamlit Cloud Secrets가 있으면 우선 적용
try:
    secrets_key = st.secrets.get("OPENAI_API_KEY", "")
    if secrets_key:
        api_key = secrets_key
except Exception:
    pass

if not api_key:
    st.error(
        "OpenAI API 키가 설정되지 않았습니다. "
        ".env 파일 또는 Streamlit Cloud Secrets에 OPENAI_API_KEY를 등록해주세요."
    )
    st.stop()

# ─────────────────────────────────────────────
# 상단 헤더 — Deploy 옆 로고
# ─────────────────────────────────────────────
_logo_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "ARASoft로고.png")
if os.path.exists(_logo_path):
    st.markdown(
        """
        <style>
        .st-emotion-cache-4xtz07 {
            height: 3rem !important;
            margin-top: 2.25rem !important;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )
    try:
        st.logo(_logo_path, size="large")
    except TypeError:
        # size 파라미터 미지원 버전
        st.logo(_logo_path)
    except AttributeError:
        # st.logo 자체 미지원 구버전 fallback
        _logo_b64 = base64.b64encode(open(_logo_path, "rb").read()).decode()
        st.markdown(
            f"""
            <style>
            [data-testid="stToolbar"]::before {{
                content: "";
                display: inline-block;
                background-image: url("data:image/png;base64,{_logo_b64}");
                background-size: contain;
                background-repeat: no-repeat;
                background-position: center;
                width: 120px;
                height: 34px;
                vertical-align: middle;
                margin-right: 8px;
            }}
            </style>
            """,
            unsafe_allow_html=True,
        )

# ─────────────────────────────────────────────
# 사이드바 — 검사 대상
# ─────────────────────────────────────────────
st.sidebar.divider()
st.sidebar.header("🎯 검사 대상")
check_audio  = st.sidebar.checkbox("🎧 음성 대본 검사", value=True)
check_screen = st.sidebar.checkbox("🖼️ 화면 텍스트 검사", value=True)
check_sb = False
use_vision_ocr_fallback = False
force_content_slides = ""
if not check_audio and not check_screen:
    st.warning("👈 검사 대상을 최소 하나 이상 체크해주세요.")
    st.stop()

# ─────────────────────────────────────────────
# 사이드바 — 문서 정보 입력
# ─────────────────────────────────────────────
st.sidebar.divider()
st.sidebar.header("📄 문서 정보 (엑셀 헤더)")
reviewer    = st.sidebar.text_input("검토자",  placeholder="예) 홍길동")
lesson_name = st.sidebar.text_input("차시명",  placeholder="예) 1강 - 파이썬 기초")
lesson_num  = st.sidebar.text_input("차시",    placeholder="예) 1")
review_date = st.sidebar.date_input("검토일자", value=datetime.date.today())

# ─────────────────────────────────────────────
# 사이드바 — 고급 설정
# ─────────────────────────────────────────────
st.sidebar.divider()
st.sidebar.header("🔧 고급 설정")

selected_model = st.sidebar.radio(
    "🤖 AI 모델 선택",
    options=["gpt-5.4", "gpt-5.4-mini"],
    index=1,
    help="gpt-5.4: 최고 정확도 / gpt-5.4-mini: 빠르고 저렴 (약 6배 저렴, 94% 성능)"
)
model_label = "GPT-5.4" if selected_model == "gpt-5.4" else "GPT-5.4 mini"

with st.sidebar.expander("화면 프레임 설정", expanded=False):
    # ★ v4: 안정화 캡처 모드 ─ 기본 ON
    stability_mode = st.checkbox(
        "🎯 안정화 캡처 (애니메이션 종료 후만 캡처)",
        value=True,
        help="영상에 페이드인·도형 그리기·텍스트 등장 등의 모션이 있을 때, "
             "애니메이션이 끝나고 화면이 안정된 시점에서만 프레임을 캡처합니다.",
    )
    if stability_mode:
        stability_min_seconds = st.slider(
            "안정 유지 최소 시간 (초)", 0.8, 4.0, 1.8, 0.2,
            help="이 시간 이상 화면 변화가 없을 때 '안정 상태'로 판정. 길게 잡으면 정확도↑(누락 가능성 약간↑)."
        )
        stability_motion_threshold = st.slider(
            "정지 판정 민감도 (낮을수록 엄격)", 0.5, 6.0, 1.5, 0.5,
            help="인접 프레임의 평균 픽셀차가 이 값 미만이면 '움직임 없음'으로 판정. "
                 "강사 미세 동작이나 3D 배경이 잡혀 캡처가 안 되면 살짝 키우세요."
        )
        stability_check_interval = st.slider(
            "내부 샘플 간격 (초)", 0.2, 1.0, 0.35, 0.05,
            help="안정성 추적용 내부 샘플링 간격. 작을수록 정밀하지만 느림."
        )
        # sample_rate / diff_threshold 는 호환용 (안정화 모드일 땐 사용 안 함)
        sample_rate = 1.5
        diff_threshold = 25.0
    else:
        stability_min_seconds = 1.8
        stability_motion_threshold = 1.5
        stability_check_interval = 0.35
        sample_rate    = st.slider("프레임 추출 간격 (초)", 0.5, 5.0, 1.5, 0.5)
        diff_threshold = st.slider("화면 변화 감지 임계값", 5.0, 50.0, 25.0, 5.0)
    # 정확도 우선: 기본 배치 2 (모델 집중도 ↑)
    batch_size     = st.selectbox("Vision API 배치 크기", [1, 2, 3, 4, 5], index=1,
                                   help="작을수록 정확도 ↑ (속도 ↓). 기본 2 권장.")

with st.sidebar.expander("음성 검사 설정", expanded=False):
    context_window     = st.slider("문맥 참고 세그먼트 수", 1, 5, 2, 1,
                                    help="교정 대상 앞뒤로 참고할 세그먼트 수")
    use_sentence_merge = st.checkbox("문장 단위 세그먼트 병합", value=True,
                                      help="Whisper가 짧게 끊어낸 세그먼트를 문장 단위로 병합 → "
                                           "문맥 오류(이중피동·어미 혼동) 검출률 향상")
    audio_batch_size   = st.slider("음성 교정 배치 크기", 10, 80, 40, 10,
                                    help="한 번에 검사할 문장 수. 작을수록 정확도 ↑")

with st.sidebar.expander("🎯 도메인 힌트 (정확도 향상)", expanded=False):
    domain_hint = st.text_area(
        "영상의 주요 용어·고유명사",
        placeholder="예) 파이썬 수업, 리스트 컴프리헨션, 재귀 함수, 객체지향",
        help="Whisper STT에 힌트로 전달되어 고유명사·전문용어 오인식을 크게 줄입니다.",
        height=80,
    )

with st.sidebar.expander("🚀 속도 설정", expanded=False):
    parallel_videos = st.slider(
        "영상 동시 처리 개수", 1, 4, 2, 1,
        help="여러 MP4를 동시에 처리합니다. OpenAI rate limit 고려 1~3 권장.",
    )
    parallel_pipelines = st.checkbox(
        "한 영상 내에서 음성·화면 검사 병렬 실행", value=True,
        help="독립적인 두 파이프라인을 동시에 실행 → 전체 시간이 긴 쪽 하나로 수렴",
    )
    audio_workers = st.slider(
        "음성 배치 병렬 호출 수", 1, 6, 3, 1,
        help="맞춤법 교정 배치를 병렬로 호출합니다. 너무 높이면 429 rate limit 발생 가능.",
    )
    screen_workers = st.slider(
        "화면 OCR 배치 병렬 호출 수", 1, 6, 4, 1,
        help="Vision API 배치 병렬 호출. Vision은 응답이 느려 병렬 효과 큼.",
    )
    jpeg_quality = st.slider(
        "JPEG 품질 (OCR 업로드)", 70, 100, 90, 5,
        help="낮추면 업로드가 빨라지지만 작은 글자 판독률↓. 90 권장.",
    )
    # 전역 인코딩 품질 적용
    set_encode_quality(jpeg_quality=jpeg_quality, max_side=1920)

with st.sidebar.expander("🛡️ 안정성 설정", expanded=False):
    use_stt_cache = st.checkbox(
        "STT 결과 캐싱 (같은 영상 재실행 시 즉시 결과)", value=True,
        help="영상 파일 내용 해시 + 도메인 힌트 조합을 키로 사용합니다. "
             "같은 영상을 다시 돌릴 때 Whisper 비용/시간 0.",
    )
    stt_chunk_workers = st.slider(
        "긴 오디오 청크 병렬 전사 수", 1, 4, 3, 1,
        help="25MB 초과 오디오가 자동으로 청크 분할될 때 병렬 전사 워커 수.",
    )
    if st.button("🗑️ STT 캐시 비우기", use_container_width=True):
        cache_root = os.path.join(tempfile.gettempdir(), "spellcheck_stt_cache")
        import shutil as _sh
        if os.path.isdir(cache_root):
            _sh.rmtree(cache_root, ignore_errors=True)
            st.success("STT 캐시 삭제 완료")
        else:
            st.info("캐시 폴더가 없습니다.")

# ─────────────────────────────────────────────
# 메인 — 문서 정보 미리보기
# ─────────────────────────────────────────────
review_date_str = review_date.strftime("%Y년 %m월 %d일") if review_date else "-"
if reviewer or lesson_name or lesson_num:
    st.markdown(f"""
    <div class="info-box">
        <h4>📄 엑셀 문서 정보</h4>
        <div class="info-grid" style="grid-template-columns:1fr 1fr 1fr 1fr;">
            <div class="info-item"><label>검토자</label><span>{reviewer or "-"}</span></div>
            <div class="info-item"><label>차시명</label><span>{lesson_name or "-"}</span></div>
            <div class="info-item"><label>차시</label><span>{lesson_num or "-"}</span></div>
            <div class="info-item"><label>검토일자</label><span>{review_date_str}</span></div>
        </div>
    </div>
    """, unsafe_allow_html=True)

# ─────────────────────────────────────────────
# 메인 — 파일 업로드 (다중 파일)
# ─────────────────────────────────────────────
st.markdown("### 🎥 동영상 업로드")
uploaded_files = st.file_uploader(
    "검사할 동영상 파일(MP4)을 업로드하세요. (여러 개 동시 업로드 가능)",
    type=["mp4"],
    accept_multiple_files=True,
    key="video_uploader"
)

if uploaded_files:
    # 업로드된 영상 미리보기 (3열 그리드, 자동 줄바꿈)
    st.markdown(f"**총 {len(uploaded_files)}개 영상 업로드됨**")
    for row_start in range(0, len(uploaded_files), 3):
        row_files = uploaded_files[row_start:row_start+3]
        cols = st.columns(3)
        for col, f in zip(cols, row_files):
            with col:
                st.caption(f"📹 {f.name}")
                st.video(f)

    st.caption(f"🤖 사용 모델: **{model_label}**  ·  🚀 영상 동시 처리: **{parallel_videos}개**  ·  "
               f"파이프라인 병렬: **{'ON' if parallel_pipelines else 'OFF'}**  ·  "
               f"🛡️ STT 캐시: **{'ON' if use_stt_cache else 'OFF'}**")

    # ── 결과 보존용 세션 상태 초기화 (한 번만) ─────────
    for _k, _v in [
        ("run_results", None),
        ("run_meta", {}),
    ]:
        if _k not in st.session_state:
            st.session_state[_k] = _v

    run_clicked = st.button(
        "🚀 맞춤법 검사 시작",
        type="primary",
        use_container_width=True,
        key="btn_run_check",
    )

    if run_clicked:
        # 새 검사 시작 → 이전 엑셀 캐시 무효화
        for k in list(st.session_state.keys()):
            if k.startswith("_excel_bytes") or k.startswith("_excel_combined_bytes"):
                del st.session_state[k]

        import time as _time
        _t0 = _time.time()

        # STT 캐시 디렉토리 (user cache temp 아래)
        stt_cache_dir = None
        if use_stt_cache:
            stt_cache_dir = os.path.join(tempfile.gettempdir(), "spellcheck_stt_cache")
            os.makedirs(stt_cache_dir, exist_ok=True)

        # ── 1) 업로드 파일을 모두 고유 임시 디렉토리에 저장 ───
        work_dir = tempfile.mkdtemp(prefix="spellcheck_")
        
        video_infos = []   # [{idx, name, filename, video_path, audio_path}]
        all_video_results = []   # try 바깥에서 미리 초기화 (예외 발생 시 NameError 방지)
        try:
            with st.spinner(f"파일 {len(uploaded_files)}개 준비 중..."):
                for file_idx, uploaded_file in enumerate(uploaded_files):
                    video_name = uploaded_file.name.rsplit(".", 1)[0]
                    video_path = os.path.join(work_dir, f"video_{file_idx}.mp4")
                    audio_path = os.path.join(work_dir, f"audio_{file_idx}.mp3")
                    with open(video_path, "wb") as f:
                        f.write(uploaded_file.getbuffer())
                    video_infos.append({
                        "idx": file_idx,
                        "name": video_name,
                        "filename": uploaded_file.name,
                        "video_path": video_path,
                        "audio_path": audio_path,
                    })

            # 자연스러운 정렬 (01_01.mp4, 01_02.mp4 순서 보장)
            def _nat_key(vi):
                return [int(t) if t.isdigit() else t.lower() for t in re.split(r'(\d+)', vi["filename"])]
            video_infos.sort(key=_nat_key)

            # ── 2) 진행 상태 UI ─────────────────────────
            overall_prog = st.progress(0.0, text=f"0 / {len(video_infos)}개 영상 처리 완료")
            status_area = st.empty()
            status_dict = {vi["idx"]: "⏳ 대기 중" for vi in video_infos}

            def _render_status():
                lines = []
                for vi in video_infos:
                    lines.append(f"- **{vi['filename']}** — {status_dict[vi['idx']]}")
                status_area.markdown("\n".join(lines))
            _render_status()

            # ── 3) 단일 영상 처리 함수 (워커 스레드에서 실행) ───
            ctx = get_script_run_ctx()

            def _process_one(vi, start_sb=0):
                if ctx and threading.current_thread().name != "MainThread":
                    add_script_run_ctx(threading.current_thread(), ctx)
                    
                try:
                    status_dict[vi["idx"]] = "🔄 처리 중..."
                    _render_status()
                    if parallel_pipelines:
                        out = run_pipeline_parallel(
                            video_path=vi["video_path"],
                            audio_path=vi["audio_path"],
                            api_key=api_key,
                            check_audio=check_audio,
                            check_screen=check_screen,
                            sample_rate=sample_rate,
                            diff_threshold=diff_threshold,
                            screen_batch_size=batch_size,
                            screen_max_workers=screen_workers,
                            model=selected_model,
                            domain_hint=domain_hint,
                            context_window=context_window,
                            use_sentence_merge=use_sentence_merge,
                            audio_batch_size=audio_batch_size,
                            audio_max_workers=audio_workers,
                            stt_chunk_workers=stt_chunk_workers,
                            stt_cache_dir=stt_cache_dir,
                            start_sb_idx=start_sb,
                            verify_pass=False,
                            slide_metadata=None,
                            stability_mode=stability_mode,
                            stability_motion_threshold=stability_motion_threshold,
                            stability_min_seconds=stability_min_seconds,
                            stability_check_interval=stability_check_interval,
                        )
                        audio_results  = out["audio"]
                        screen_results = out["screen"]
                        errors = out.get("errors", {})
                        stt_cached = out.get("stt_cached", False)
                        last_sb = out.get("last_sb_idx", start_sb)
                        slide_map = out.get("slide_map", [])
                        slide_coverage = out.get("slide_coverage", {})
                    else:
                        audio_results, screen_results = [], []
                        errors = {}
                        stt_cached = False
                        last_sb = start_sb
                        slide_map = []
                        slide_coverage = {}
                        if check_screen:
                            try:
                                frames = extract_and_filter_frames(
                                    vi["video_path"], sample_rate, diff_threshold,
                                    stability_mode=stability_mode,
                                    stability_motion_threshold=stability_motion_threshold,
                                    stability_min_seconds=stability_min_seconds,
                                    stability_check_interval=stability_check_interval,
                                )
                                if frames:
                                    screen_results, slide_map, last_sb = spell_check_frames(
                                        frames, api_key,
                                        batch_size=batch_size,
                                        model=selected_model,
                                        max_workers=screen_workers,
                                        start_sb_idx=start_sb,
                                        verify_pass=False,
                                        slide_metadata=None,
                                    )
                            except Exception as e:
                                errors["screen"] = str(e)
                        if check_audio:
                            try:
                                if extract_audio(vi["video_path"], vi["audio_path"]):
                                    segments = transcribe_audio(
                                        vi["audio_path"], api_key,
                                        domain_hint=domain_hint,
                                        max_workers=stt_chunk_workers,
                                    )
                                    if segments:
                                        audio_results = spell_check_segments(
                                            segments, api_key,
                                            context_window=context_window,
                                            model=selected_model,
                                            use_sentence_merge=use_sentence_merge,
                                            batch_size=audio_batch_size,
                                            max_workers=audio_workers,
                                        )
                                        audio_results = attach_frames_to_audio_results(
                                            audio_results, vi["video_path"])
                                else:
                                    errors["audio"] = "오디오 추출 실패"
                            except Exception as e:
                                errors["audio"] = str(e)

                    err_note = ""
                    if errors:
                        err_note = " ⚠️ " + ", ".join(f"{k}: {v[:40]}" for k, v in errors.items())
                    cache_badge = " ⚡캐시" if stt_cached else ""
                    status_dict[vi["idx"]] = (
                        f"✅ 완료{cache_badge} — 🎧 {len(audio_results)}건 / "
                        f"🖼️ {len(screen_results)}건{err_note}"
                    )

                    return {
                        "idx": vi["idx"],
                        "name": vi["name"],
                        "filename": vi["filename"],
                        "audio": audio_results,
                        "screen": screen_results,
                        "errors": errors,
                        "stt_cached": stt_cached,
                        "last_sb_idx": last_sb,
                        "slide_map": slide_map,
                        "slide_coverage": slide_coverage,
                        "is_fatal": False,
                    }
                except Exception as e:
                    status_dict[vi["idx"]] = f"❌ 실패: {str(e)[:100]}"
                    return {
                        "idx": vi["idx"],
                        "name": vi["name"],
                        "filename": vi["filename"],
                        "audio": [],
                        "screen": [],
                        "errors": {"fatal": str(e)},
                        "stt_cached": False,
                        "last_sb_idx": start_sb,
                        "slide_map": [],
                        "slide_coverage": {},
                        "is_fatal": True,
                    }

            # ── 4) 영상 간 병렬 실행 ──────────────────────
            results_by_idx = {}
            done_count = 0
            if parallel_videos <= 1 or len(video_infos) == 1 or check_sb:
                cur_sb = 0
                for vi in video_infos:
                    res = _process_one(vi, cur_sb)
                    results_by_idx[vi["idx"]] = res
                    cur_sb = res.get("last_sb_idx", cur_sb)
                    done_count += 1
                    overall_prog.progress(
                        done_count / len(video_infos),
                        text=f"{done_count} / {len(video_infos)}개 영상 처리 완료",
                    )
                    _render_status()
            else:
                with ThreadPoolExecutor(max_workers=parallel_videos) as ex:
                    futures = {ex.submit(_process_one, vi): vi["idx"] for vi in video_infos}
                    for fut in as_completed(futures):
                        idx = futures[fut]
                        try:
                            results_by_idx[idx] = fut.result()
                        except Exception as e:
                            results_by_idx[idx] = {
                                "idx": idx, "name": "-", "filename": "-",
                                "audio": [], "screen": [],
                                "errors": {"fatal": str(e)},
                                "stt_cached": False, "is_fatal": True,
                                "slide_map": [], "slide_coverage": {},
                            }
                        done_count += 1
                        overall_prog.progress(
                            done_count / len(video_infos),
                            text=f"{done_count} / {len(video_infos)}개 영상 처리 완료",
                        )
                        _render_status()

            # 원본 순서대로 정렬
            all_video_results = [results_by_idx[vi["idx"]] for vi in video_infos]

            elapsed = _time.time() - _t0
            overall_prog.progress(1.0, text=f"✅ 전체 완료 ({elapsed:.1f}초)")

            success_count = sum(1 for r in all_video_results if not r.get("is_fatal"))
            fail_count = len(all_video_results) - success_count
            cache_hits = sum(1 for r in all_video_results if r.get("stt_cached"))

            summary_msg = (
                f"🎉 {success_count}/{len(all_video_results)}개 성공 · "
                f"소요 시간 **{elapsed:.1f}초**"
            )
            if cache_hits:
                summary_msg += f" · ⚡ STT 캐시 적중 {cache_hits}개"
            if fail_count:
                summary_msg += f" · ❌ 실패 {fail_count}개"
                st.error(summary_msg)
                # 실패 영상의 상세 에러 표시
                for r in all_video_results:
                    if r.get("is_fatal"):
                        st.caption(f"❌ **{r['filename']}**: {r['errors'].get('fatal', '알 수 없는 오류')}")
            else:
                st.success(summary_msg)

        finally:
            # ── 임시 작업 디렉토리 확실히 정리 ────────────
            import shutil as _shutil
            try:
                _shutil.rmtree(work_dir, ignore_errors=True)
            except Exception:
                pass

        # 결과가 하나도 없으면 (파일 준비 실패 등) 여기서 중단
        if not all_video_results:
            st.stop()

        # ★ 결과를 세션 상태에 보존 → 다운로드 클릭으로 rerun되어도 유지됨
        st.session_state.run_results = all_video_results
        st.session_state.run_meta = {
            "reviewer":     reviewer,
            "lesson_name":  lesson_name,
            "lesson_num":   lesson_num,
            "review_date":  review_date,
            "check_audio":  check_audio,
            "check_screen": check_screen,
            "check_sb":     check_sb,
        }

        # ★ 결과 저장 완료 후 즉시 rerun → 이제부터 단일 렌더 경로(session_state 기반)만 사용
        # (이렇게 하면 첫 표시와 다운로드 후 표시가 완전히 동일한 코드를 거치므로 변수 차이가 없음)
        st.rerun()

# ── 결과 렌더링 (검사 시작 블록 밖, 업로드 상태와 독립) ────
# 다운로드 버튼을 눌러 rerun이 발생해도 session_state에 보존된 결과로 다시 그립니다.
if st.session_state.get("run_results"):
    all_video_results = st.session_state.run_results
    _meta = st.session_state.run_meta
    reviewer    = _meta.get("reviewer", "")
    lesson_name = _meta.get("lesson_name", "")
    lesson_num  = _meta.get("lesson_num", "")
    review_date = _meta.get("review_date")
    check_audio = _meta.get("check_audio", True)
    check_screen = _meta.get("check_screen", True)
    check_sb     = False

    if True:

        # ── 전체 결과 출력 ────────────────────────────
        st.divider()

        # 결과 헤더 + 리셋 버튼 (결과가 있을 때만 노출)
        head_col1, head_col2 = st.columns([5, 1])
        with head_col1:
            st.subheader("📋 교정 결과")
        with head_col2:
            if st.button(
                "🔄 결과 리셋",
                use_container_width=True,
                key="btn_reset_results",
                help="현재 표시 중인 검사 결과를 화면에서 지웁니다. 업로드 파일은 유지됩니다.",
            ):
                st.session_state.run_results = None
                st.session_state.run_meta = {}
                for k in list(st.session_state.keys()):
                    if k.startswith("_excel_bytes") or k.startswith("_excel_combined_bytes"):
                        del st.session_state[k]
                st.rerun()

        # 영상별 탭
        score_rows = build_score_rows(all_video_results)
        if score_rows:
            render_score_summary_card(score_rows)
            with st.expander("영상별 점수 보기", expanded=len(score_rows) <= 4):
                st.dataframe(
                    pd.DataFrame([
                        {
                            "영상명": r["video"],
                            "점수": r["score"],
                            "등급": r["grade"],
                            "음성 오류": r["audio_count"],
                            "화면 오류": r["screen_count"],
                            "총 오류": r["total_count"],
                        }
                        for r in score_rows
                    ]),
                    use_container_width=True,
                    hide_index=True,
                )

        video_tab_labels = [
            f"📹 {r['name']} ({len(r['audio'])+len(r['screen'])}건)"
            for r in all_video_results
        ]
        video_tabs = st.tabs(video_tab_labels)

        for v_tab, vr in zip(video_tabs, all_video_results):
            with v_tab:
                audio_results  = vr["audio"]
                screen_results = vr["screen"]
                slide_map_data = []
                coverage_data  = {}
                score_info = calculate_score(audio_results, screen_results)

                # 요약 지표
                c1, c2, c3, c4 = st.columns(4)
                c1.metric("전체 교정 건수", f"{len(audio_results)+len(screen_results)}건")
                c2.metric("🎧 음성 대본",   f"{len(audio_results)}건")
                c3.metric("🖼️ 화면 텍스트", f"{len(screen_results)}건")
                c4.metric("🧮 품질 점수", f"{score_info['score']}점", score_info["grade"])
                st.markdown("")

                # 음성/화면 탭
                inner_labels = []
                if check_audio:  inner_labels.append(f"🎧 음성 대본 ({len(audio_results)}건)")
                if check_screen: inner_labels.append(f"🖼️ 화면 텍스트 ({len(screen_results)}건)")
                inner_tabs = st.tabs(inner_labels)
                inner_idx = 0

                if check_audio:
                    with inner_tabs[inner_idx]:
                        st.markdown("##### 음성에서 발견된 맞춤법·오타 오류")
                        render_result_cards(audio_results, badge_type="audio")
                    inner_idx += 1

                if check_screen:
                    with inner_tabs[inner_idx]:
                        st.markdown("##### 화면 텍스트에서 발견된 맞춤법·오타 오류")
                        render_result_cards(screen_results, badge_type="screen")
                    inner_idx += 1

                # 엑셀 다운로드 — 캐시해서 rerun 시 재생성 비용 회피
                if not vr.get("is_fatal"):
                    st.divider()
                    excel_cache_key = f"_excel_bytes_screen_score_only_v1::{vr['name']}"
                    if excel_cache_key not in st.session_state:
                        st.session_state[excel_cache_key] = make_excel(
                            audio_results, screen_results,
                            reviewer, lesson_name, lesson_num, review_date,
                        )
                    excel_bytes = st.session_state[excel_cache_key]
                    base = vr["name"]
                    fname = f"{lesson_num}차시_{base}_교정결과.xlsx" if lesson_name else f"{base}_교정결과.xlsx"
                    st.download_button(
                        label=f"📥 {vr['filename']} 교정 결과 다운로드",
                        data=excel_bytes,
                        file_name=fname,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                        type="primary",
                        key=f"dl_{vr['name']}"
                    )
                    sheet_desc = "음성 대본 / 화면 텍스트 / 통합"
                    st.info(f"💡 엑셀 파일에는 **{sheet_desc}** 시트가 포함됩니다.")

        # ── 전체 통합 다운로드 (영상 2개 이상일 때만 표시) ──
        if len(all_video_results) >= 2:
            total_cnt = sum(len(vr["audio"]) + len(vr["screen"]) for vr in all_video_results)
            if score_rows:
                st.divider()
                st.markdown("### 📦 전체 통합 다운로드")
                st.caption(
                    f"총 {len(all_video_results)}개 영상 · {total_cnt}건 교정 항목"
                    + " 을 하나의 엑셀 파일로 다운로드합니다."
                )
                combined_cache_key = "_excel_combined_bytes_screen_score_only_v1"
                if combined_cache_key not in st.session_state:
                    st.session_state[combined_cache_key] = make_excel_combined(
                        all_video_results, reviewer, lesson_name, lesson_num, review_date,
                    )
                combined_bytes = st.session_state[combined_cache_key]
                fname_combined = f"{lesson_name}_전체통합_교정결과.xlsx" if lesson_name else "전체통합_교정결과.xlsx"
                st.download_button(
                    label="📥 전체 통합 엑셀 다운로드 (.xlsx)",
                    data=combined_bytes,
                    file_name=fname_combined,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    type="primary",
                    key="dl_combined"
                )
                struct_desc = "[영상명_음성] [영상명_화면]"
                struct_desc += " 시트 × 영상 수 + [📋 전체통합] 시트"
                st.info(f"💡 통합 파일 구조: {struct_desc}")

